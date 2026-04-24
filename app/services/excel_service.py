import json
import logging
import os
import socket
import time
from collections.abc import Callable
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import ENCABEZADOS, HOJA_REGISTROS, get_consolidado_path, get_excel_path, normalizar_sede_archivo
from app.services.settings_service import get_settings

logger = logging.getLogger(__name__)


class ArchivoCajaOcupadoError(Exception):
    pass


_LOCK_TTL = 60  # segundos — lock más antiguo que esto se considera huérfano


SECTION_PREFIXES = {
    "caja": "Caja",
    "plataformas": "Plataformas",
    "gastos": "Gastos",
    "bonos": "Bonos",
    "prestamos": "Prestamos",
    "movimientos": "Movimientos",
    "contadores": "Contadores",
    "cuadre": "Cuadre",
}

BONOS_HEADERS = [
    "fecha",
    "hora",
    "cliente",
    "valor_bono",
    "fecha_hora_registro",
]

PRESTAMOS_HEADERS = [
    "fecha",
    "hora",
    "persona",
    "tipo_movimiento",
    "valor_movimiento",
    "fecha_hora_registro",
]

MOVIMIENTOS_HEADERS = [
    "fecha",
    "hora",
    "tipo_movimiento",
    "concepto",
    "valor",
    "observacion",
    "fecha_hora_registro",
]

PLATAFORMAS_HEADERS = [
    "fecha",
    "venta_practisistemas",
    "venta_deportivas",
    "total_plataformas",
    "fecha_hora_registro",
]

CUADRE_HEADERS = [
    "fecha",
    "fecha_inicio_periodo",
    "base_anterior",
    "total_contadores",
    "total_practisistemas",
    "total_deportivas",
    "total_bonos",
    "total_gastos",
    "total_prestamos_salida",
    "total_prestamos_entrada",
    "neto_prestamos",
    "total_mov_ingresos",
    "total_mov_salidas",
    "neto_movimientos",
    "caja_teorica",
    "caja_fisica",
    "diferencia",
    "base_nueva",
    "fecha_hora_registro",
]

CONTADORES_HEADERS = [
    "fecha",
    "item_id",
    "nombre",
    "denominacion",
    "entradas",
    "salidas",
    "jackpot",
    "yield_actual",
    "ref_entradas",
    "ref_salidas",
    "ref_jackpot",
    "yield_referencia",
    "produccion_pre_reset",
    "observacion",
    "resultado_monetario",
    "fecha_hora_registro",
]


def _path_modulo(modulo: str, year: int) -> Path:
    """Devuelve la ruta del Excel según el módulo.
    Cuadre y Contadores van a Consolidado_{sede}_{año}.xlsx; el resto a Contadores_{sede}_{año}.xlsx.
    """
    if modulo in {"cuadre", "contadores"}:
        return get_consolidado_path(year)
    return get_excel_path(year)


def _abrir_o_crear_workbook(path: Path):
    if path.exists():
        try:
            return load_workbook(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "El libro de Excel esta ocupado por otro proceso. Intenta guardar de nuevo en unos segundos."
            ) from exc
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _normalizar_nombre_hoja(nombre: str | None) -> str:
    base = (nombre or "").strip() or "Principal"
    invalidos = set(':\\/?*[]')
    limpio = "".join("_" if ch in invalidos else ch for ch in base)
    return limpio[:31].strip() or "Principal"


def _obtener_nombre_sede() -> str:
    from app.config import _get_active_dir_and_sede
    _, sede = _get_active_dir_and_sede()
    return _normalizar_nombre_hoja(sede)


def _obtener_nombre_hoja_seccion(seccion: str) -> str:
    prefijo = SECTION_PREFIXES.get(seccion, seccion.title())
    return _normalizar_nombre_hoja(f"{prefijo}{_obtener_nombre_sede()}")


def obtener_hojas_activas() -> dict[str, str]:
    enabled = get_settings().get("enabled_modules") or ["caja"]
    return {
        modulo: _obtener_nombre_hoja_seccion(modulo)
        for modulo in enabled
        if modulo in SECTION_PREFIXES
    }


def _nombres_unicos(nombres: list[str]) -> list[str]:
    return [nombre for nombre in dict.fromkeys(nombres) if nombre]


def _nombres_legacy_caja() -> list[str]:
    return _nombres_unicos([
        _obtener_nombre_sede(),
        HOJA_REGISTROS,
    ])


def _nombres_lectura_modulo(modulo: str) -> list[str]:
    if modulo == "caja":
        return _nombres_unicos([
            _obtener_nombre_hoja_seccion("caja"),
            *_nombres_legacy_caja(),
        ])
    if modulo == "gastos":
        return _nombres_unicos([
            _obtener_nombre_hoja_seccion("gastos"),
            _obtener_nombre_hoja_seccion("caja"),
            *_nombres_legacy_caja(),
        ])
    if modulo == "plataformas":
        return _nombres_unicos([_obtener_nombre_hoja_seccion("plataformas")])
    return _nombres_unicos([_obtener_nombre_hoja_seccion(modulo)])


def _fila_es_modulo(modulo: str, row) -> bool:
    if modulo in {"bonos", "prestamos", "movimientos", "plataformas", "contadores", "cuadre"}:
        return True
    tipo = row[1]
    if modulo == "caja":
        return tipo not in {"gasto", "bono"}
    if modulo == "gastos":
        return tipo == "gasto"
    return False


def _lock_metadata() -> bytes:
    return json.dumps({
        "ts": time.time(),
        "pid": os.getpid(),
        "host": socket.gethostname(),
    }).encode()


def _lock_host(lock_path: Path) -> str:
    try:
        return json.loads(lock_path.read_bytes()).get("host", "")
    except Exception:
        return ""


def _lock_age_seconds(lock_path: Path) -> float:
    try:
        data = json.loads(lock_path.read_bytes())
        ts = float(data.get("ts"))
        return max(0.0, time.time() - ts)
    except Exception:
        try:
            return max(0.0, time.time() - lock_path.stat().st_mtime)
        except OSError:
            return 0.0


@contextmanager
def _bloqueo_escritura(path: Path):
    # Capa 1: archivo propietario que Excel crea al abrir el xlsx (Windows).
    # Dropbox lo sincroniza, así que detecta si otro equipo lo tiene abierto en Excel.
    owner_path = path.parent / f"~${path.name}"
    if owner_path.exists():
        raise ArchivoCajaOcupadoError(
            "El libro esta siendo usado por Excel en este equipo o en otro equipo de la red. "
            "Cierra el archivo en Excel antes de guardar."
        )

    # Capa 2: archivo .lock con metadata (host, pid, timestamp).
    # Dropbox lo sincroniza, bloqueando guardados simultáneos entre distintos equipos.
    # Si el lock tiene más de _LOCK_TTL segundos se considera huérfano y se elimina automáticamente.
    lock_path = path.with_suffix(path.suffix + ".lock")
    fd = None
    lock_creado = False
    try:
        for intento in range(2):
            try:
                fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_RDWR)
                os.write(fd, _lock_metadata())
                lock_creado = True
                break
            except FileExistsError:
                age = _lock_age_seconds(lock_path)
                if age > _LOCK_TTL and intento == 0:
                    logger.warning("Se detectó un lock huérfano en %s (%.1fs); se eliminará y se reintentará.", lock_path, age)
                    lock_path.unlink(missing_ok=True)
                    continue  # reintento único tras limpiar lock huérfano
                host = _lock_host(lock_path)
                suffix = f" desde el equipo '{host}'" if host else ""
                logger.info("Lock activo detectado para %s%s", path, suffix)
                raise ArchivoCajaOcupadoError(
                    f"Otro usuario esta guardando en este momento{suffix}. "
                    "Vuelve a intentarlo en unos segundos."
                )
            except PermissionError as exc:
                logger.warning("No se pudo crear el lock de escritura para %s por permisos/sincronización: %s", path, exc)
                raise ArchivoCajaOcupadoError(
                    "No se pudo guardar porque el libro esta abierto o sincronizandose. "
                    "Cierra Excel y vuelve a intentarlo."
                ) from exc
            except OSError as exc:
                logger.warning("No se pudo bloquear el libro %s para escritura: %s", path, exc)
                raise ArchivoCajaOcupadoError(
                    "No se pudo bloquear el libro para guardar. Vuelve a intentarlo en unos segundos."
                ) from exc
        yield
    finally:
        if fd is not None:
            os.close(fd)
        if lock_creado and lock_path.exists():
            try:
                lock_path.unlink(missing_ok=True)
            except Exception as exc:
                logger.warning("No se pudo limpiar el lock %s al finalizar la escritura: %s", lock_path, exc)


@contextmanager
def _abrir_workbook_lectura(path: Path):
    wb = load_workbook(path, read_only=True)
    try:
        yield wb
    finally:
        wb.close()


def _escribir_encabezados(ws, modulo: str):
    if modulo == "bonos":
        headers = BONOS_HEADERS
    elif modulo == "prestamos":
        headers = PRESTAMOS_HEADERS
    elif modulo == "movimientos":
        headers = MOVIMIENTOS_HEADERS
    elif modulo == "plataformas":
        headers = PLATAFORMAS_HEADERS
    elif modulo == "contadores":
        headers = CONTADORES_HEADERS
    elif modulo == "cuadre":
        headers = CUADRE_HEADERS
    else:
        headers = ENCABEZADOS
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    if modulo == "bonos":
        widths = {"A": 12, "B": 12, "C": 28, "D": 14, "E": 22}
    elif modulo == "prestamos":
        widths = {"A": 12, "B": 12, "C": 28, "D": 18, "E": 14, "F": 22}
    elif modulo == "movimientos":
        widths = {"A": 12, "B": 12, "C": 16, "D": 28, "E": 14, "F": 28, "G": 22}
    elif modulo == "plataformas":
        widths = {"A": 14, "B": 18, "C": 18, "D": 18, "E": 22}
    elif modulo == "contadores":
        widths = {
            "A": 14, "B": 20, "C": 26, "D": 14, "E": 12,
            "F": 12, "G": 12, "H": 14, "I": 12, "J": 12,
            "K": 12, "L": 14, "M": 18, "N": 30, "O": 18, "P": 22,
        }
    elif modulo == "cuadre":
        widths = {
            "A": 14, "B": 14, "C": 16, "D": 16, "E": 16,
            "F": 16, "G": 14, "H": 14, "I": 16, "J": 16,
            "K": 14, "L": 16, "M": 16, "N": 14, "O": 16,
            "P": 16, "Q": 16, "R": 16, "S": 22,
        }
    else:
        widths = {"A": 14, "B": 16, "C": 22, "D": 14, "E": 10, "F": 14, "G": 14, "H": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _asegurar_presentacion_hoja(ws, modulo: str) -> None:
    if ws.freeze_panes != "A2":
        ws.freeze_panes = "A2"

    if modulo == "bonos":
        widths = {"A": 12, "B": 12, "C": 28, "D": 14, "E": 22}
    elif modulo == "prestamos":
        widths = {"A": 12, "B": 12, "C": 28, "D": 18, "E": 14, "F": 22}
    elif modulo == "movimientos":
        widths = {"A": 12, "B": 12, "C": 16, "D": 28, "E": 14, "F": 28, "G": 22}
    elif modulo == "plataformas":
        widths = {"A": 14, "B": 18, "C": 18, "D": 18, "E": 22}
    elif modulo == "contadores":
        widths = {
            "A": 14, "B": 20, "C": 26, "D": 14, "E": 12,
            "F": 12, "G": 12, "H": 14, "I": 12, "J": 12,
            "K": 12, "L": 14, "M": 18, "N": 30, "O": 18, "P": 22,
        }
    elif modulo == "cuadre":
        widths = {
            "A": 14, "B": 14, "C": 16, "D": 16, "E": 16,
            "F": 16, "G": 14, "H": 14, "I": 16, "J": 16,
            "K": 14, "L": 16, "M": 16, "N": 14, "O": 16,
            "P": 16, "Q": 16, "R": 16, "S": 22,
        }
    else:
        widths = {"A": 14, "B": 16, "C": 22, "D": 14, "E": 10, "F": 14, "G": 14, "H": 22}

    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _marcar_celda_activa(ws, row_num: int) -> None:
    try:
        ws.parent.active = ws.parent.index(ws)
    except (AttributeError, ValueError, TypeError):
        pass


def _hoja_tiene_datos(ws) -> bool:
    return ws.max_row > 1


def _puede_migrar_desde_registros_diarios(wb) -> bool:
    if HOJA_REGISTROS not in wb.sheetnames:
        return False

    for name in wb.sheetnames:
        if name == HOJA_REGISTROS:
            continue
        if _hoja_tiene_datos(wb[name]):
            return False
    return True


def _obtener_hojas_para_lectura(wb, modulo: str):
    return [wb[name] for name in _nombres_lectura_modulo(modulo) if name in wb.sheetnames]


def _asegurar_hoja(wb, modulo: str):
    hoja_destino = _obtener_nombre_hoja_seccion(modulo)
    if hoja_destino in wb.sheetnames:
        ws = wb[hoja_destino]
        _asegurar_presentacion_hoja(ws, modulo)
        return ws

    if modulo == "caja":
        for legacy_name in _nombres_legacy_caja():
            if legacy_name == HOJA_REGISTROS:
                continue
            if legacy_name in wb.sheetnames:
                wb[legacy_name].title = hoja_destino
                ws = wb[hoja_destino]
                _asegurar_presentacion_hoja(ws, modulo)
                return ws

        if _puede_migrar_desde_registros_diarios(wb):
            ws_legacy = wb[HOJA_REGISTROS]
            ws_legacy.title = hoja_destino
            _asegurar_presentacion_hoja(ws_legacy, modulo)
            return ws_legacy

    ws = wb.create_sheet(hoja_destino)
    _escribir_encabezados(ws, modulo)
    return ws


def _actualizar_encabezados_prestamos(ws) -> None:
    headers_actuales = [str(ws.cell(1, idx).value or "").strip().lower() for idx in range(1, 7)]
    if headers_actuales == PRESTAMOS_HEADERS:
        return
    for idx, header in enumerate(PRESTAMOS_HEADERS, start=1):
        ws.cell(1, idx).value = header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = {"A": 12, "B": 12, "C": 28, "D": 18, "E": 14, "F": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _iterar_filas_fecha(hojas, fecha_objetivo: date):
    for ws in hojas:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            cell_date = row[0]
            if isinstance(cell_date, datetime):
                cell_date = cell_date.date()
            if isinstance(cell_date, date) and cell_date == fecha_objetivo:
                yield row


def _formatear_filas_recientes(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        7: "#,##0",
        8: "HH:mm AM/PM",
    })


def _formatear_filas_por_columnas(ws, cantidad_filas: int, formatos: dict[int, str]) -> None:
    if cantidad_filas <= 0:
        return

    last_row = ws.max_row
    start_row = last_row - cantidad_filas + 1
    for row_num in range(start_row, last_row + 1):
        for col, formato in formatos.items():
            ws.cell(row_num, col).number_format = formato


def _formatear_fila(ws, row_num: int, formatos: dict[int, str]) -> None:
    for col, formato in formatos.items():
        ws.cell(row_num, col).number_format = formato


def _formatear_filas_recientes_bonos(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        2: "HH:mm AM/PM",
        4: "#,##0",
        5: "HH:mm AM/PM",
    })


def _formatear_filas_recientes_prestamos(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        2: "HH:mm AM/PM",
        5: "#,##0",
        6: "HH:mm AM/PM",
    })


def _formatear_filas_recientes_movimientos(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        2: "HH:mm AM/PM",
        5: "#,##0",
        7: "HH:mm AM/PM",
    })


def _formatear_filas_recientes_cuadre(ws, cantidad_filas: int) -> None:
    formatos = {1: "DD-MM-YYYY", 2: "DD-MM-YYYY", 19: "HH:mm AM/PM"}
    formatos.update({col: "#,##0" for col in range(3, 19)})
    _formatear_filas_por_columnas(ws, cantidad_filas, formatos)


def _formatear_filas_recientes_contadores(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        15: "#,##0",
        16: "HH:mm AM/PM",
    })


def _formatear_filas_recientes_plataformas(ws, cantidad_filas: int) -> None:
    _formatear_filas_por_columnas(ws, cantidad_filas, {
        1: "DD-MM-YYYY",
        2: "#,##0",
        3: "#,##0",
        4: "#,##0",
        5: "HH:mm AM/PM",
    })


def fecha_existe_modulo(modulo: str, fecha: date, year: int) -> bool:
    path = _path_modulo(modulo, year)
    if not path.exists():
        return False

    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, modulo)
        if not hojas:
            return False

        for row in _iterar_filas_fecha(hojas, fecha):
            if _fila_es_modulo(modulo, row):
                return True

    return False


def guardar_filas_modulo(modulo: str, filas: list, year: int, reemplazar_fecha: date | None = None) -> None:
    path = _path_modulo(modulo, year)
    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        if reemplazar_fecha is not None:
            _eliminar_fecha_modulo_en_workbook(wb, modulo, reemplazar_fecha)

        if filas:
            ws = _asegurar_hoja(wb, modulo)
            for fila in filas:
                ws.append(fila)
            if modulo == "bonos":
                _formatear_filas_recientes_bonos(ws, len(filas))
            elif modulo == "prestamos":
                _formatear_filas_recientes_prestamos(ws, len(filas))
            elif modulo == "movimientos":
                _formatear_filas_recientes_movimientos(ws, len(filas))
            elif modulo == "contadores":
                _formatear_filas_recientes_contadores(ws, len(filas))
            elif modulo == "cuadre":
                _formatear_filas_recientes_cuadre(ws, len(filas))
            elif modulo == "plataformas":
                _formatear_filas_recientes_plataformas(ws, len(filas))
            else:
                _formatear_filas_recientes(ws, len(filas))
            _marcar_celda_activa(ws, ws.max_row)

        try:
            wb.save(path)
        except PermissionError as exc:
            logger.warning("Excel ocupado al guardar %s en %s: %s", modulo, path, exc)
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()


def _eliminar_fecha_modulo_en_workbook(wb, modulo: str, fecha: date) -> int:
    total_borradas = 0
    for nombre in _nombres_lectura_modulo(modulo):
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        filas_a_borrar = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            cell_date = row[0]
            if isinstance(cell_date, datetime):
                cell_date = cell_date.date()
            if isinstance(cell_date, date) and cell_date == fecha and _fila_es_modulo(modulo, row):
                filas_a_borrar.append(i)

        for row_num in reversed(filas_a_borrar):
            ws.delete_rows(row_num)
        total_borradas += len(filas_a_borrar)
    return total_borradas


def eliminar_fecha_modulo(modulo: str, fecha: date, year: int) -> int:
    path = _path_modulo(modulo, year)
    if not path.exists():
        return 0

    with _bloqueo_escritura(path):
        try:
            wb = load_workbook(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo actualizar el libro porque esta ocupado. Intenta nuevamente."
            ) from exc

        total_borradas = _eliminar_fecha_modulo_en_workbook(wb, modulo, fecha)
        try:
            wb.save(path)
        except PermissionError as exc:
            logger.warning("Excel ocupado al eliminar %s de %s en %s: %s", fecha, modulo, path, exc)
            raise ArchivoCajaOcupadoError(
                "No se pudo actualizar el libro porque esta ocupado. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
        return total_borradas


def obtener_fechas_modulo_año(modulo: str, year: int) -> list[str]:
    path = _path_modulo(modulo, year)
    if not path.exists():
        return []

    fechas = set()
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, modulo)
        if not hojas:
            return []

        for ws in hojas:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                cell_date = row[0]
                if isinstance(cell_date, datetime):
                    cell_date = cell_date.date()
                if isinstance(cell_date, date) and cell_date.year == year and _fila_es_modulo(modulo, row):
                    fechas.add(str(cell_date))

    return sorted(fechas)


def obtener_datos_caja_fecha(fecha: date, year: int) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    resultado = {
        "billetes": {},
        "total_monedas": 0,
        "billetes_viejos": 0,
    }
    found = False

    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "caja")
        if not hojas:
            return None

        for row in _iterar_filas_fecha(hojas, fecha):
            if not _fila_es_modulo("caja", row):
                continue
            found = True
            tipo, concepto, subtotal = row[1], row[2], row[6] or 0
            if tipo == "billete":
                denom, cantidad = row[3], row[4]
                if denom is not None:
                    resultado["billetes"][str(int(denom))] = int(cantidad or 0)
            elif tipo == "manual":
                if concepto == "Total monedas":
                    resultado["total_monedas"] = subtotal
                elif concepto == "Billetes viejos":
                    resultado["billetes_viejos"] = subtotal

    return resultado if found else None


def obtener_datos_plataformas_fecha(fecha: date, year: int) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "plataformas")
        if not hojas:
            return None
        for row in _iterar_filas_fecha(hojas, fecha):
            practi = float(row[1] or 0)
            deport = float(row[2] or 0)
            total = float(row[3] or 0)
            ts = row[4]
            return {
                "venta_practisistemas": practi,
                "venta_deportivas": deport,
                "total_plataformas": total,
                "fecha_hora_registro": ts.isoformat() if isinstance(ts, datetime) else str(ts or ""),
            }
    return None


def obtener_items_modulo_fecha(modulo: str, fecha: date, year: int) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    items = []
    total = 0
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, modulo)
        if not hojas:
            return None

        for row in _iterar_filas_fecha(hojas, fecha):
            if not _fila_es_modulo(modulo, row):
                continue
            concepto = row[2] or ""
            valor = row[6] or 0
            items.append({"concepto": concepto, "valor": valor})
            total += valor

    if not items:
        return None
    return {"items": items, "total": total}


def obtener_ultima_fecha_modulo(modulo: str, year: int):
    path = _path_modulo(modulo, year)
    if not path.exists():
        return None

    ultima = None
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, modulo)
        if not hojas:
            return None

        for ws in hojas:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                if modulo not in {"bonos", "prestamos"} and not _fila_es_modulo(modulo, row):
                    continue
                cell_date = row[0]
                if isinstance(cell_date, datetime):
                    cell_date = cell_date.date()
                if not isinstance(cell_date, date):
                    continue
                if ultima is None or cell_date > ultima:
                    ultima = cell_date

    return ultima


def obtener_ultima_fecha_modulo_global(modulo: str, years_back: int = 5):
    hoy = date.today()
    ultima = None
    for year in range(hoy.year, hoy.year - years_back, -1):
        candidata = obtener_ultima_fecha_modulo(modulo, year)
        if candidata is None:
            continue
        if ultima is None or candidata > ultima:
            ultima = candidata
    return ultima


def obtener_datos_contadores_fecha(fecha: date, year: int) -> dict:
    """Devuelve {item_id: {...}} con los valores guardados para la fecha dada."""
    path = get_consolidado_path(year)
    if not path.exists():
        return {}
    result = {}
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "contadores")
        if not hojas:
            return {}
        for row in _iterar_filas_fecha(hojas, fecha):
            data = _parsear_fila_contadores(row)
            item_id = data["item_id"]
            if not item_id:
                continue
            result[item_id] = {
                "entradas": data["entradas"],
                "salidas": data["salidas"],
                "jackpot": data["jackpot"],
                "yield_actual": data["yield_actual"],
                "yield_referencia": data["yield_referencia"],
                "produccion_pre_reset": data["produccion_pre_reset"],
                "resultado_unidades": data["yield_actual"] - data["yield_referencia"],
                "resultado_monetario": data["resultado_monetario"],
                "observacion": data["observacion"],
                "ref_entradas": data["ref_entradas"],
                "ref_salidas": data["ref_salidas"],
                "ref_jackpot": data["ref_jackpot"],
                "fecha_hora_registro": data["fecha_hora_registro"],
            }
    return result


def obtener_historial_contadores(hasta_fecha: date) -> list[dict]:
    """Devuelve todos los registros de contadores anteriores a hasta_fecha, en todos los años disponibles.
    Si la fila tiene referencia crítica embebida, emite un evento adicional de tipo referencia_critica
    en la misma fecha (con orden posterior al registro normal) para que _iter_referencias_previas
    lo use como referencia vigente.
    """
    eventos = []
    for path in _obtener_paths_consolidado_sede():
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "contadores")
            for ws in hojas:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    cell_date = row[0]
                    if isinstance(cell_date, datetime):
                        cell_date = cell_date.date()
                    if not isinstance(cell_date, date):
                        continue
                    if cell_date >= hasta_fecha:
                        continue
                    item_id = str(row[1] or "").strip()
                    if not item_id:
                        continue
                    fecha_str = str(cell_date)
                    data = _parsear_fila_contadores(row)
                    observacion = data["observacion"]
                    eventos.append({
                        "tipo": "registro",
                        "fecha": fecha_str,
                        "item_id": item_id,
                        "entradas": data["entradas"],
                        "salidas": data["salidas"],
                        "jackpot": data["jackpot"],
                        "yield": data["yield_actual"],
                        "observacion": observacion,
                    })
                    if data["ref_entradas"] is not None:
                        ref_e = int(data["ref_entradas"] or 0)
                        ref_s = int(data["ref_salidas"] or 0)
                        ref_j = int(data["ref_jackpot"] or 0)
                        eventos.append({
                            "tipo": "referencia_critica",
                            "fecha": fecha_str,
                            "item_id": item_id,
                            "entradas": ref_e,
                            "salidas": ref_s,
                            "jackpot": ref_j,
                            "yield": ref_e - ref_s - ref_j,
                            "observacion": observacion,
                        })
    return eventos


def _obtener_paths_sede(prefijo: str, path_factory: Callable[[int], Path]) -> list[Path]:
    settings = get_settings()
    data_dir = Path(settings.get("data_dir") or ".")
    sede = normalizar_sede_archivo(settings.get("sede"))
    patron = f"{prefijo}_{sede}_*.xlsx"
    paths = sorted(data_dir.glob(patron))
    actual = path_factory(date.today().year)
    if actual not in paths:
        paths.append(actual)
    return paths


def _obtener_paths_excel_sede() -> list[Path]:
    return _obtener_paths_sede("Contadores", get_excel_path)


def _obtener_paths_consolidado_sede() -> list[Path]:
    return _obtener_paths_sede("Consolidado", get_consolidado_path)


def _parsear_fila_contadores(row) -> dict:
    # fecha,item_id,nombre,denom,entradas,salidas,jackpot,yield_actual,
    # ref_entradas,ref_salidas,ref_jackpot,yield_referencia,produccion_pre_reset,
    # observacion,resultado_monetario,fecha_hora_registro
    return {
        "item_id": str(row[1] or "").strip(),
        "entradas": int(row[4] or 0),
        "salidas": int(row[5] or 0),
        "jackpot": int(row[6] or 0),
        "yield_actual": int(row[7] or 0),
        "yield_referencia": int(row[11] or 0),
        "produccion_pre_reset": int(row[12] or 0),
        "observacion": str(row[13] or ""),
        "resultado_monetario": float(row[14] or 0),
        "ref_entradas": int(row[8]) if row[8] is not None else None,
        "ref_salidas": int(row[9]) if row[9] is not None else None,
        "ref_jackpot": int(row[10]) if row[10] is not None else None,
        "fecha_hora_registro": row[15].isoformat() if isinstance(row[15], datetime) else str(row[15] or ""),
    }


def _leer_movimientos_prestamos_desde_hoja(
    ws,
    *,
    persona_norm: str | None = None,
    fecha_objetivo: date | None = None,
    fechas_objetivo: set[date] | None = None,
) -> list[dict]:
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    schema_nuevo = headers[:6] == PRESTAMOS_HEADERS
    registros = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        cell_date = row[0]
        if isinstance(cell_date, datetime):
            cell_date = cell_date.date()
        if not isinstance(cell_date, date):
            continue
        if fecha_objetivo and cell_date != fecha_objetivo:
            continue
        if fechas_objetivo and cell_date not in fechas_objetivo:
            continue

        if schema_nuevo:
            hora = row[1]
            persona = str(row[2] or "").strip()
            tipo = str(row[3] or "prestamo").strip().lower() or "prestamo"
            valor = float(row[4] or 0)
            timestamp = row[5]
        else:
            hora = None
            persona = str(row[1] or "").strip()
            tipo = "prestamo"
            valor = float(row[2] or 0)
            timestamp = row[3]

        if not persona:
            continue
        if persona_norm and persona.strip().lower() != persona_norm:
            continue
        if isinstance(hora, datetime):
            hora_texto = hora.strftime("%I:%M %p")
        elif hora:
            hora_texto = str(hora)
        elif isinstance(timestamp, datetime):
            hora_texto = timestamp.strftime("%I:%M %p")
        else:
            hora_texto = ""

        registros.append({
            "sheet_row": idx,
            "fecha": cell_date.isoformat(),
            "fecha_display": cell_date.strftime("%d-%m-%Y"),
            "hora_display": hora_texto,
            "persona": persona,
            "tipo_movimiento": "pago" if tipo == "pago" else "prestamo",
            "valor": abs(valor),
            "fecha_hora_registro": timestamp.isoformat() if isinstance(timestamp, datetime) else str(timestamp or ""),
        })
    return registros


def _resumen_prestamos_desde_registros(registros: list[dict]) -> dict:
    total_prestado = sum(float(item["valor"] or 0) for item in registros if item["tipo_movimiento"] == "prestamo")
    total_pagado = sum(float(item["valor"] or 0) for item in registros if item["tipo_movimiento"] == "pago")
    saldo_pendiente = total_prestado - total_pagado
    return {
        "items": registros,
        "total_prestado": total_prestado,
        "total_pagado": total_pagado,
        "saldo_pendiente": saldo_pendiente,
    }

def guardar_bono_registro(fecha: date, cliente: str, valor: float, timestamp: datetime) -> float:
    fila = [fecha, timestamp, cliente, valor, timestamp]
    path = get_excel_path(fecha.year)
    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "bonos")
        total_dia = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_date = row[0]
            if isinstance(cell_date, datetime):
                cell_date = cell_date.date()
            if isinstance(cell_date, date) and cell_date == fecha:
                total_dia += float(row[3] or 0)

        ws.append(fila)
        _formatear_filas_recientes_bonos(ws, 1)
        _marcar_celda_activa(ws, ws.max_row)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return total_dia + float(valor or 0)


def guardar_prestamo_registro(
    fecha: date,
    persona: str,
    tipo_movimiento: str,
    valor: float,
    timestamp: datetime,
) -> None:
    fila = [fecha, timestamp, persona, tipo_movimiento, valor, timestamp]
    path = get_excel_path(fecha.year)
    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "prestamos")
        _actualizar_encabezados_prestamos(ws)
        ws.append(fila)
        _formatear_filas_recientes_prestamos(ws, 1)
        _marcar_celda_activa(ws, ws.max_row)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()


def guardar_movimiento_registro(
    fecha: date,
    tipo_movimiento: str,
    concepto: str,
    valor: float,
    observacion: str,
    timestamp: datetime,
) -> dict:
    fila = [fecha, timestamp, tipo_movimiento, concepto, valor, observacion, timestamp]
    path = get_excel_path(fecha.year)
    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "movimientos")
        ws.append(fila)
        _formatear_filas_recientes_movimientos(ws, 1)
        _marcar_celda_activa(ws, ws.max_row)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    registros = obtener_movimientos_fecha(fecha, fecha.year)
    total_ingresos = sum(float(item["valor"] or 0) for item in registros if item["tipo_movimiento"] == "ingreso")
    total_salidas = sum(float(item["valor"] or 0) for item in registros if item["tipo_movimiento"] == "salida")
    return {
        "total_ingresos": total_ingresos,
        "total_salidas": total_salidas,
        "neto": total_ingresos - total_salidas,
    }


def obtener_bonos_fecha(fecha: date, year: int) -> list[dict]:
    path = get_excel_path(year)
    if not path.exists():
        return []

    registros = []
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "bonos")
        if not hojas:
            return []

        for ws in hojas:
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:
                    continue
                cell_date = row[0]
                if isinstance(cell_date, datetime):
                    cell_date = cell_date.date()
                if not (isinstance(cell_date, date) and cell_date == fecha):
                    continue
                hora = row[1]
                if isinstance(hora, datetime):
                    hora_texto = hora.strftime("%I:%M %p")
                else:
                    hora_texto = str(hora or "")
                valor = float(row[3] or 0)
                registros.append({
                    "sheet_row": idx,
                    "fecha": cell_date.isoformat(),
                    "fecha_display": cell_date.strftime("%d-%m-%Y"),
                    "hora_display": hora_texto,
                    "cliente": str(row[2] or ""),
                    "valor": valor,
                    "fecha_hora_registro": row[4].isoformat() if isinstance(row[4], datetime) else str(row[4] or ""),
                })
    registros.sort(key=lambda item: item["fecha_hora_registro"] or "")
    return registros


def obtener_prestamos_fecha(fecha: date, year: int) -> list[dict]:
    path = get_excel_path(year)
    if not path.exists():
        return []

    registros = []
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "prestamos")
        if not hojas:
            return []
        for ws in hojas:
            registros.extend(_leer_movimientos_prestamos_desde_hoja(ws, fecha_objetivo=fecha))
    return registros


def obtener_movimientos_fecha(fecha: date, year: int) -> list[dict]:
    path = get_excel_path(year)
    if not path.exists():
        return []

    registros = []
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "movimientos")
        if not hojas:
            return []

        for ws in hojas:
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:
                    continue
                cell_date = row[0]
                if isinstance(cell_date, datetime):
                    cell_date = cell_date.date()
                if not (isinstance(cell_date, date) and cell_date == fecha):
                    continue
                hora = row[1]
                if isinstance(hora, datetime):
                    hora_texto = hora.strftime("%I:%M %p")
                else:
                    hora_texto = str(hora or "")
                registros.append({
                    "sheet_row": idx,
                    "fecha": cell_date.isoformat(),
                    "fecha_display": cell_date.strftime("%d-%m-%Y"),
                    "hora_display": hora_texto,
                    "tipo_movimiento": str(row[2] or "").strip().lower() or "salida",
                    "concepto": str(row[3] or "").strip(),
                    "valor": float(row[4] or 0),
                    "observacion": str(row[5] or "").strip(),
                    "fecha_hora_registro": row[6].isoformat() if isinstance(row[6], datetime) else str(row[6] or ""),
                })
    registros.sort(key=lambda item: item["fecha_hora_registro"] or "", reverse=True)
    return registros


def _construir_ciclo_activo_prestamos(registros: list[dict], fecha_ref: date | None = None) -> list[dict]:
    from datetime import date as _date
    fecha_ref_iso = (fecha_ref or _date.today()).isoformat()

    registros = sorted(registros, key=lambda item: (item["fecha"], item["fecha_hora_registro"] or ""))

    saldos: dict[str, float] = {}
    ciclo_inicio: dict[str, int] = {}
    ciclo_cierre: dict[str, tuple[int, str]] = {}

    for i, item in enumerate(registros):
        persona_key = item["persona"].strip().lower()
        saldo_prev = saldos.get(persona_key, 0.0)
        if abs(saldo_prev) < 0.01 and item["tipo_movimiento"] == "prestamo":
            ciclo_inicio[persona_key] = i
        saldo = saldo_prev
        if item["tipo_movimiento"] == "pago":
            saldo -= float(item["valor"] or 0)
        else:
            saldo += float(item["valor"] or 0)
        item["saldo_pendiente"] = round(saldo, 2)
        saldos[persona_key] = saldo
        if abs(saldo) < 0.01:
            ciclo_cierre[persona_key] = (i, item["fecha"])

    resultado = []
    for i, item in enumerate(registros):
        persona_key = item["persona"].strip().lower()
        inicio = ciclo_inicio.get(persona_key, 0)
        saldo_final = saldos.get(persona_key, 0.0)
        if abs(saldo_final) >= 0.01:
            if i >= inicio:
                resultado.append(item)
        else:
            cierre = ciclo_cierre.get(persona_key)
            if cierre and cierre[1] == fecha_ref_iso and inicio <= i <= cierre[0]:
                resultado.append(item)
    return resultado


def obtener_movimientos_prestamos(persona: str | None = None, fecha_hasta: date | None = None) -> list[dict]:
    persona_norm = persona.strip().lower() if persona else None
    registros = []
    for path in _obtener_paths_excel_sede():
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "prestamos")
            for ws in hojas:
                registros.extend(_leer_movimientos_prestamos_desde_hoja(ws, persona_norm=persona_norm))
    if fecha_hasta:
        fecha_tope = fecha_hasta.isoformat()
        registros = [item for item in registros if item["fecha"] <= fecha_tope]
    return _construir_ciclo_activo_prestamos(registros, fecha_ref=fecha_hasta)


def obtener_resumen_prestamos(persona: str | None = None, fecha_hasta: date | None = None) -> dict:
    return _resumen_prestamos_desde_registros(obtener_movimientos_prestamos(persona=persona, fecha_hasta=fecha_hasta))


def obtener_movimientos_prestamos_super_admin(fecha_objetivo: date) -> dict:
    registros = []
    fecha_iso = fecha_objetivo.isoformat()
    for path in _obtener_paths_excel_sede():
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "prestamos")
            for ws in hojas:
                registros.extend(_leer_movimientos_prestamos_desde_hoja(ws))

    registros.sort(key=lambda item: (item["fecha"], item["fecha_hora_registro"] or ""))

    saldos_por_persona: dict[str, float] = {}
    items_del_dia: list[dict] = []

    for item in registros:
        if item["fecha"] > fecha_iso:
            break
        persona_key = item["persona"].strip().lower()
        saldo = saldos_por_persona.get(persona_key, 0.0)
        valor = float(item["valor"] or 0)
        if item["tipo_movimiento"] == "pago":
            saldo -= valor
        else:
            saldo += valor
        saldo = round(saldo, 2)
        saldos_por_persona[persona_key] = saldo
        if item["fecha"] == fecha_iso:
            item_dia = dict(item)
            item_dia["saldo_pendiente"] = saldo
            items_del_dia.append(item_dia)

    deuda_total_activa = round(sum(max(saldo, 0.0) for saldo in saldos_por_persona.values()), 2)
    return {
        "items": items_del_dia,
        "saldos_por_persona": {k: round(v, 2) for k, v in saldos_por_persona.items()},
        "deuda_total_activa": deuda_total_activa,
    }

def obtener_prestamos_modulo(fecha_objetivo: date) -> dict:
    """Items del ciclo activo (con saldo_pendiente=0 en día de cierre) + saldos_por_persona acumulados."""
    fecha_iso = fecha_objetivo.isoformat()
    registros = []
    for path in _obtener_paths_excel_sede():
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "prestamos")
            for ws in hojas:
                registros.extend(_leer_movimientos_prestamos_desde_hoja(ws))

    registros_hasta = [
        item for item in registros
        if item["fecha"] <= fecha_iso
    ]
    registros_hasta.sort(key=lambda item: (item["fecha"], item["fecha_hora_registro"] or ""))

    saldos_por_persona: dict[str, float] = {}
    for item in registros_hasta:
        persona_key = item["persona"].strip().lower()
        saldo = saldos_por_persona.get(persona_key, 0.0)
        if item["tipo_movimiento"] == "pago":
            saldo -= float(item["valor"] or 0)
        else:
            saldo += float(item["valor"] or 0)
        saldos_por_persona[persona_key] = round(saldo, 2)

    deuda_total_activa = round(sum(max(s, 0.0) for s in saldos_por_persona.values()), 2)
    items_ciclo = _construir_ciclo_activo_prestamos(registros_hasta, fecha_ref=fecha_objetivo)

    return {
        "items": items_ciclo,
        "saldos_por_persona": {k: round(v, 2) for k, v in saldos_por_persona.items()},
        "deuda_total_activa": deuda_total_activa,
    }


def obtener_prestamos_raw_fecha(fecha: date, year: int) -> list[dict]:
    """Todos los movimientos de préstamos de una fecha, sin filtro de ciclo activo."""
    path = get_excel_path(year)
    if not path.exists():
        return []
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "prestamos")
        if not hojas:
            return []
        registros = []
        for ws in hojas:
            for r in _leer_movimientos_prestamos_desde_hoja(ws):
                if r["fecha"] == fecha.isoformat():
                    registros.append(r)
    return registros


def _parsear_fila_cuadre(row) -> dict:
    # fecha, fecha_inicio_periodo, base_anterior, total_contadores, total_practisistemas,
    # total_deportivas, total_bonos, total_gastos, total_prestamos_salida, total_prestamos_entrada,
    # neto_prestamos, total_mov_ingresos, total_mov_salidas, neto_movimientos,
    # caja_teorica, caja_fisica, diferencia, base_nueva, fecha_hora_registro
    fecha_val = row[0]
    if isinstance(fecha_val, datetime):
        fecha_val = fecha_val.date()
    fecha_inicio = row[1]
    if isinstance(fecha_inicio, datetime):
        fecha_inicio = fecha_inicio.date()
    return {
        "fecha": str(fecha_val) if fecha_val else "",
        "fecha_inicio_periodo": str(fecha_inicio) if fecha_inicio else "",
        "base_anterior": float(row[2] or 0),
        "total_contadores": float(row[3] or 0),
        "total_practisistemas": float(row[4] or 0),
        "total_deportivas": float(row[5] or 0),
        "total_bonos": float(row[6] or 0),
        "total_gastos": float(row[7] or 0),
        "total_prestamos_salida": float(row[8] or 0),
        "total_prestamos_entrada": float(row[9] or 0),
        "neto_prestamos": float(row[10] or 0),
        "total_mov_ingresos": float(row[11] or 0),
        "total_mov_salidas": float(row[12] or 0),
        "neto_movimientos": float(row[13] or 0),
        "caja_teorica": float(row[14] or 0),
        "caja_fisica": float(row[15] or 0),
        "diferencia": float(row[16] or 0),
        "base_nueva": float(row[17] or 0),
        "fecha_hora_registro": (
            row[18].isoformat() if isinstance(row[18], datetime) else str(row[18] or "")
        ),
    }


def obtener_datos_cuadre_fecha(fecha: date, year: int) -> dict | None:
    path = _path_modulo("cuadre", year)
    if not path.exists():
        return None
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "cuadre")
        if not hojas:
            return None
        for row in _iterar_filas_fecha(hojas, fecha):
            return _parsear_fila_cuadre(row)
    return None


def obtener_cuadre_que_contiene_fecha(fecha_op: date) -> dict | None:
    candidato = None
    cierre_candidato = None

    for year in {fecha_op.year, fecha_op.year + 1}:
        path = _path_modulo("cuadre", year)
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "cuadre")
            if not hojas:
                continue
            for ws in hojas:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None or row[1] is None:
                        continue
                    cuadre = _parsear_fila_cuadre(row)
                    try:
                        inicio = date.fromisoformat(cuadre["fecha_inicio_periodo"])
                        cierre = date.fromisoformat(cuadre["fecha"])
                    except ValueError:
                        continue
                    if not (inicio <= fecha_op <= cierre):
                        continue
                    if cierre_candidato is None or cierre > cierre_candidato:
                        candidato = cuadre
                        cierre_candidato = cierre

    return candidato


def obtener_siguiente_cuadre(fecha_cierre: date) -> dict | None:
    candidato = None
    cierre_candidato = None

    for year in {fecha_cierre.year, fecha_cierre.year + 1}:
        path = _path_modulo("cuadre", year)
        if not path.exists():
            continue
        with _abrir_workbook_lectura(path) as wb:
            hojas = _obtener_hojas_para_lectura(wb, "cuadre")
            if not hojas:
                continue
            for ws in hojas:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    try:
                        cuadre = _parsear_fila_cuadre(row)
                        cierre = date.fromisoformat(cuadre["fecha"])
                    except (ValueError, TypeError, KeyError):
                        continue
                    if cierre <= fecha_cierre:
                        continue
                    if cierre_candidato is None or cierre < cierre_candidato:
                        candidato = cuadre
                        cierre_candidato = cierre

    return candidato


def _ts_matches(cell_val, ts_str: str) -> bool:
    if isinstance(cell_val, datetime):
        return cell_val.isoformat() == ts_str
    return str(cell_val or "").strip() == ts_str


def _obtener_hoja_existente_modulo(wb, modulo: str):
    for nombre in _nombres_lectura_modulo(modulo):
        if nombre in wb.sheetnames:
            ws = wb[nombre]
            _asegurar_presentacion_hoja(ws, modulo)
            return ws
    return None


def obtener_gastos_fecha(fecha: date, year: int) -> list[dict]:
    path = get_excel_path(year)
    if not path.exists():
        return []

    registros = []
    with _abrir_workbook_lectura(path) as wb:
        hojas = _obtener_hojas_para_lectura(wb, "gastos")
        if not hojas:
            return []
        for ws in hojas:
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is None:
                    continue
                cell_date = row[0]
                if isinstance(cell_date, datetime):
                    cell_date = cell_date.date()
                if not (isinstance(cell_date, date) and cell_date == fecha):
                    continue
                if str(row[1] or "").strip().lower() != "gasto":
                    continue
                ts = row[7]
                registros.append({
                    "sheet_row": idx,
                    "fecha": cell_date.isoformat(),
                    "fecha_display": cell_date.strftime("%d-%m-%Y"),
                    "hora_display": ts.strftime("%I:%M %p") if isinstance(ts, datetime) else "",
                    "concepto": str(row[2] or "").strip(),
                    "valor": float(row[6] or 0),
                    "fecha_hora_registro": ts.isoformat() if isinstance(ts, datetime) else str(ts or ""),
                })
    registros.sort(key=lambda item: item["fecha_hora_registro"] or "", reverse=True)
    return registros


def actualizar_bono_por_ts(fecha: date, year: int, ts_str: str, cliente: str, valor: float, new_ts: datetime) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "bonos")
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[4].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.cell(target, 3).value = cliente
        ws.cell(target, 4).value = valor
        ws.cell(target, 5).value = new_ts
        _formatear_fila(ws, target, {1: "DD-MM-YYYY", 2: "HH:mm AM/PM", 4: "#,##0", 5: "HH:mm AM/PM"})
        _marcar_celda_activa(ws, target)
        total_dia = sum(
            float(r[3] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if (r[0].date() if isinstance(r[0], datetime) else r[0]) == fecha
        )
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return {"cliente": cliente, "valor": valor, "total_dia": total_dia}


def eliminar_bono_por_ts(fecha: date, year: int, ts_str: str) -> float | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "bonos")
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[4].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.delete_rows(target)
        _marcar_celda_activa(ws, ws.max_row)
        total_dia = sum(
            float(r[3] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if (r[0].date() if isinstance(r[0], datetime) else r[0]) == fecha
        )
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return total_dia


def actualizar_gasto_por_ts(fecha: date, year: int, ts_str: str, concepto: str, valor: float, new_ts: datetime) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _obtener_hoja_existente_modulo(wb, "gastos")
        if ws is None:
            wb.close()
            return None
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if str(row[1].value or "").strip().lower() == "gasto" and _ts_matches(row[7].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.cell(target, 3).value = concepto
        ws.cell(target, 7).value = valor
        ws.cell(target, 8).value = new_ts
        _formatear_fila(ws, target, {1: "DD-MM-YYYY", 7: "#,##0", 8: "HH:mm AM/PM"})
        _marcar_celda_activa(ws, target)
        total_dia = sum(
            float(r[6] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if str(r[1] or "").strip().lower() == "gasto"
            and (r[0].date() if isinstance(r[0], datetime) else r[0]) == fecha
        )
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return {"concepto": concepto, "valor": valor, "total_dia": total_dia}


def eliminar_gasto_por_ts(fecha: date, year: int, ts_str: str) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _obtener_hoja_existente_modulo(wb, "gastos")
        if ws is None:
            wb.close()
            return None
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if str(row[1].value or "").strip().lower() == "gasto" and _ts_matches(row[7].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.delete_rows(target)
        _marcar_celda_activa(ws, ws.max_row)
        total_dia = sum(
            float(r[6] or 0)
            for r in ws.iter_rows(min_row=2, values_only=True)
            if str(r[1] or "").strip().lower() == "gasto"
            and (r[0].date() if isinstance(r[0], datetime) else r[0]) == fecha
        )
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return {"total_dia": total_dia}


def actualizar_prestamo_por_ts(fecha: date, year: int, ts_str: str, persona: str, tipo: str, valor: float, new_ts: datetime) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "prestamos")
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[5].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.cell(target, 3).value = persona
        ws.cell(target, 4).value = tipo
        ws.cell(target, 5).value = valor
        ws.cell(target, 6).value = new_ts
        _formatear_fila(ws, target, {1: "DD-MM-YYYY", 2: "HH:mm AM/PM", 5: "#,##0", 6: "HH:mm AM/PM"})
        _marcar_celda_activa(ws, target)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return obtener_resumen_prestamos(persona=persona, fecha_hasta=fecha)


def eliminar_prestamo_por_ts(fecha: date, year: int, ts_str: str) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "prestamos")
        target = None
        persona = ""
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[5].value, ts_str):
                target = idx
                persona = str(row[2].value or "").strip()
                break
        if target is None:
            wb.close()
            return None
        ws.delete_rows(target)
        _marcar_celda_activa(ws, ws.max_row)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    return obtener_resumen_prestamos(persona=persona, fecha_hasta=fecha)


def actualizar_movimiento_por_ts(fecha: date, year: int, ts_str: str, tipo: str, concepto: str, valor: float, observacion: str, new_ts: datetime) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "movimientos")
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[6].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.cell(target, 3).value = tipo
        ws.cell(target, 4).value = concepto
        ws.cell(target, 5).value = valor
        ws.cell(target, 6).value = observacion
        ws.cell(target, 7).value = new_ts
        _formatear_fila(ws, target, {1: "DD-MM-YYYY", 2: "HH:mm AM/PM", 5: "#,##0", 7: "HH:mm AM/PM"})
        _marcar_celda_activa(ws, target)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    items = obtener_movimientos_fecha(fecha, year)
    total_ingresos = sum(float(i["valor"] or 0) for i in items if i["tipo_movimiento"] == "ingreso")
    total_salidas = sum(float(i["valor"] or 0) for i in items if i["tipo_movimiento"] == "salida")
    return {"items": items, "total_ingresos": total_ingresos, "total_salidas": total_salidas, "neto": total_ingresos - total_salidas}


def eliminar_movimiento_por_ts(fecha: date, year: int, ts_str: str) -> dict | None:
    path = get_excel_path(year)
    if not path.exists():
        return None

    with _bloqueo_escritura(path):
        wb = _abrir_o_crear_workbook(path)
        ws = _asegurar_hoja(wb, "movimientos")
        target = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if _ts_matches(row[6].value, ts_str):
                target = idx
                break
        if target is None:
            wb.close()
            return None
        ws.delete_rows(target)
        _marcar_celda_activa(ws, ws.max_row)
        try:
            wb.save(path)
        except PermissionError as exc:
            raise ArchivoCajaOcupadoError(
                "No se pudo guardar porque el libro esta siendo usado por otro proceso. Intenta nuevamente."
            ) from exc
        finally:
            wb.close()
    items = obtener_movimientos_fecha(fecha, year)
    total_ingresos = sum(float(i["valor"] or 0) for i in items if i["tipo_movimiento"] == "ingreso")
    total_salidas = sum(float(i["valor"] or 0) for i in items if i["tipo_movimiento"] == "salida")
    return {"items": items, "total_ingresos": total_ingresos, "total_salidas": total_salidas, "neto": total_ingresos - total_salidas}
