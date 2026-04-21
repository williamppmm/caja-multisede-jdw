"""
Servicio de lectura de referencias externas para el módulo Plataformas.

Lee dos archivos Excel externos (Practisistemas y Bet/Deportivas) y extrae
el valor de venta para una fecha y sede determinadas. El resultado se usa
exclusivamente como referencia visual — nunca se guarda en los xlsx propios.

Cache por (ruta_archivo, fecha_iso, encabezado, mtime) para no releer el
workbook en cada petición dentro de la misma sesión.
"""

from __future__ import annotations

import threading
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.settings_service import get_plataformas_ref_config

# ── Cache ─────────────────────────────────────────────────────────────────────

_cache: dict[tuple, dict] = {}
_cache_lock = threading.Lock()

_PRACTI_FILENAME = "Ventas_dia_Practisistemas.xlsx"
_BET_FILENAME = "Ventas_dia_Bet.xlsm"

# Hojas donde buscar en cada archivo
_PRACTI_SHEET = "Resumen"
_BET_SHEET = "xDias"

# Nombres de la columna de fecha en cada hoja
_PRACTI_FECHA_COL = "Fecha"
_BET_FECHA_COL = "FECHA"


# ── Helpers internos ──────────────────────────────────────────────────────────

def _mtime(path: Path) -> float | None:
    try:
        return path.stat().st_mtime
    except OSError:
        return None


def _normalizar_valor(raw: Any) -> tuple[str, float | None]:
    """Devuelve (status, valor_float).

    status puede ser: ok | sin_dato | vacio
    """
    if raw is None or str(raw).strip() == "":
        return "vacio", None
    texto = str(raw).strip().lower()
    if texto in {"sin datos", "sin dato", "-", "n/a", "nd"}:
        return "sin_dato", None
    try:
        valor = float(str(raw).replace(",", ".").replace(" ", ""))
        return "ok", valor
    except (ValueError, TypeError):
        return "sin_dato", None


def _fecha_coincide(cell_value: Any, fecha: date) -> bool:
    """Compara el valor de una celda contra la fecha buscada."""
    if cell_value is None:
        return False
    if isinstance(cell_value, (datetime,)):
        return cell_value.date() == fecha
    if isinstance(cell_value, date):
        return cell_value == fecha
    # Texto: intentar parsear formatos comunes
    texto = str(cell_value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt).date() == fecha
        except ValueError:
            continue
    return False


def _coerce_fecha(raw: Any) -> date | None:
    """Convierte un valor devuelto por otros servicios a `date`.

    `resolver_periodo_operativo()` puede devolver fechas como objetos `date`
    o como strings ISO. Este helper normaliza ambas variantes.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    texto = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def _leer_valor(path: Path, sheet_name: str, fecha_col: str,
                header: str, fecha: date, keep_vba: bool = False) -> dict:
    """Abre el workbook, encuentra la fila de la fecha y extrae el valor
    de la columna `header`. Devuelve un dict con status y valor."""
    if not path.exists():
        return {"status": "archivo_no_encontrado", "valor": None, "header": header}

    try:
        wb = load_workbook(path, read_only=True, keep_vba=keep_vba, data_only=True)
    except Exception as exc:
        return {"status": "error", "valor": None, "header": header, "detalle": str(exc)}

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"status": "hoja_no_encontrada", "valor": None, "header": header}

    ws = wb[sheet_name]

    # Leer encabezados de la primera fila
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return {"status": "hoja_no_encontrada", "valor": None, "header": header}

    # Índice de columna fecha y columna del encabezado buscado (base 0)
    fecha_idx: int | None = None
    valor_idx: int | None = None
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        nombre = str(cell).strip()
        if nombre == fecha_col:
            fecha_idx = i
        if nombre == header:
            valor_idx = i

    if fecha_idx is None:
        wb.close()
        return {"status": "hoja_no_encontrada", "valor": None, "header": header}

    if valor_idx is None:
        wb.close()
        return {"status": "columna_no_encontrada", "valor": None, "header": header}

    # Buscar fila con la fecha
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(fecha_idx, valor_idx):
            continue
        if _fecha_coincide(row[fecha_idx], fecha):
            status, valor = _normalizar_valor(row[valor_idx])
            wb.close()
            return {"status": status, "valor": valor, "header": header}

    wb.close()
    return {"status": "fecha_no_encontrada", "valor": None, "header": header}


# ── API pública ───────────────────────────────────────────────────────────────

def obtener_referencias(fecha: date) -> dict:
    """Devuelve las referencias de Practisistemas y Bet para la fecha y sede activas.

    Respuesta:
    {
        "ok": True,
        "sede": "Barbacoas",
        "practisistemas": {"status": "ok", "valor": 24000, "header": "localbarbacoas"},
        "deportivas":     {"status": "sin_dato", "valor": None, "header": "Barbacoas"},
    }
    """
    cfg = get_plataformas_ref_config()
    practi_path_str = cfg["practi_path"]
    bet_path_str = cfg["bet_path"]
    practi_header = cfg["practi_header"]
    bet_header = cfg["bet_header"]

    # Resolver rutas de archivo
    practi_file = (Path(practi_path_str) / _PRACTI_FILENAME) if practi_path_str else None
    bet_file = (Path(bet_path_str) / _BET_FILENAME) if bet_path_str else None

    fecha_iso = fecha.isoformat()

    def _consultar(archivo: Path | None, sheet: str, fecha_col: str,
                   header: str, keep_vba: bool, fuente: str) -> dict:
        if not archivo:
            return {"status": "sin_ruta", "valor": None, "header": header}
        if not header:
            return {"status": "sin_mapeo", "valor": None, "header": header}

        cache_key = (str(archivo), fecha_iso, header, _mtime(archivo))
        with _cache_lock:
            if cache_key in _cache:
                return _cache[cache_key]

        resultado = _leer_valor(archivo, sheet, fecha_col, header, fecha, keep_vba)

        with _cache_lock:
            _cache[cache_key] = resultado
        return resultado

    try:
        from app.services import cuadre_service
        periodo_info = cuadre_service.resolver_periodo_operativo(fecha)
        fechas_periodo_raw = periodo_info.get("periodo") or [fecha]
        fechas_periodo = [fecha_ref for item in fechas_periodo_raw if (fecha_ref := _coerce_fecha(item)) is not None]
        if not fechas_periodo:
            fechas_periodo = [fecha]
    except Exception:
        fechas_periodo = [fecha]

    def _consultar_fecha(archivo: Path | None, sheet: str, fecha_col: str,
                        header: str, keep_vba: bool, fecha_ref: date) -> dict:
        if not archivo:
            return {"status": "sin_ruta", "valor": None, "header": header}
        if not header:
            return {"status": "sin_mapeo", "valor": None, "header": header}

        cache_key = (str(archivo), fecha_ref.isoformat(), header, _mtime(archivo))
        with _cache_lock:
            if cache_key in _cache:
                return _cache[cache_key]

        resultado = _leer_valor(archivo, sheet, fecha_col, header, fecha_ref, keep_vba)

        with _cache_lock:
            _cache[cache_key] = resultado
        return resultado

    def _acumular(archivo: Path | None, sheet: str, fecha_col: str,
                  header: str, keep_vba: bool) -> dict:
        resultados = [
            _consultar_fecha(archivo, sheet, fecha_col, header, keep_vba, fecha_ref)
            for fecha_ref in fechas_periodo
        ]

        total = 0.0
        ok_count = 0
        ultimo_fallo = None
        for resultado in resultados:
            status = resultado.get("status")
            if status == "ok" and resultado.get("valor") is not None:
                total += float(resultado["valor"] or 0)
                ok_count += 1
            else:
                ultimo_fallo = resultado

        base = {
            "header": header,
            "desde": fechas_periodo[0].isoformat() if fechas_periodo else fecha.isoformat(),
            "hasta": fechas_periodo[-1].isoformat() if fechas_periodo else fecha.isoformat(),
            "dias": len(fechas_periodo),
        }

        if ok_count == len(resultados) and ok_count > 0:
            return {**base, "status": "ok", "valor": round(total, 2)}
        if ok_count > 0:
            return {**base, "status": "parcial", "valor": round(total, 2), "detalle": ultimo_fallo.get("status") if ultimo_fallo else ""}
        if ultimo_fallo:
            return {**base, **ultimo_fallo}
        return {**base, "status": "sin_dato", "valor": None}

    practi = _acumular(practi_file, _PRACTI_SHEET, _PRACTI_FECHA_COL, practi_header, False)
    deportivas = _acumular(bet_file, _BET_SHEET, _BET_FECHA_COL, bet_header, True)

    # Sede activa para contexto en la respuesta
    from app.services.settings_service import get_active_site, get_settings
    settings = get_settings()
    if settings.get("super_admin_mode"):
        site = get_active_site()
        sede = site["sede"] if site else ""
    else:
        sede = settings.get("sede", "")

    return {
        "ok": True,
        "sede": sede,
        "desde": fechas_periodo[0].isoformat() if fechas_periodo else fecha.isoformat(),
        "hasta": fechas_periodo[-1].isoformat() if fechas_periodo else fecha.isoformat(),
        "dias": len(fechas_periodo),
        "practisistemas": practi,
        "deportivas": deportivas,
    }
