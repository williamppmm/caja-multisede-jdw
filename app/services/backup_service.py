import json
import logging
import shutil
import tempfile
import threading
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

RETENTION_DAYS = 3
DELAY_SECONDS = 600        # 10 minutos post-arranque
REPEAT_SECONDS = 4 * 3600  # re-verificar cada 4 horas

_JSON_FILES = ["contadores_items.json", "startup_state.json"]

_backup_lock = threading.Lock()
_scheduled_lock = threading.Lock()
_backup_scheduled = False


# ── Helpers internos ──────────────────────────────────────────────────────────

def _es_fecha(nombre: str) -> bool:
    try:
        date.fromisoformat(nombre)
        return True
    except ValueError:
        return False


def _resolver_archivos_fuente(data_dir: Path) -> list[Path]:
    """Devuelve los archivos a respaldar de una sede: xlsx con patrón conocido + json fijos."""
    archivos: list[Path] = []
    for patron in ("Contadores_*.xlsx", "Consolidado_*.xlsx"):
        archivos.extend(sorted(data_dir.glob(patron)))
    for nombre in _JSON_FILES:
        p = data_dir / nombre
        if p.exists():
            archivos.append(p)
    return archivos


def _validar(path: Path) -> bool:
    """Verifica que el archivo sea legible. xlsx → openpyxl; json → parse.
    Nota: Workbook no es context manager, se cierra manualmente con try/finally.
    """
    try:
        if path.suffix == ".xlsx":
            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                _ = wb.sheetnames
            finally:
                wb.close()
        elif path.suffix == ".json":
            with open(path, encoding="utf-8") as fh:
                json.load(fh)
        return True
    except Exception:
        return False


def _leer_manifest(carpeta_dia: Path) -> dict | None:
    manifest_path = carpeta_dia / "manifest.json"
    if not manifest_path.exists():
        return None
    try:
        with open(manifest_path, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return None


def _backup_dia_completo(carpeta_dia: Path) -> bool:
    """Un respaldo del día es completo si existe manifest.json con valido=True."""
    manifest = _leer_manifest(carpeta_dia)
    return bool(manifest and manifest.get("valido", False))


def _escribir_manifest(carpeta_dia: Path, manifest: dict) -> None:
    tmp = carpeta_dia / "manifest.json.tmp"
    dst = carpeta_dia / "manifest.json"
    try:
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(manifest, fh, indent=2, ensure_ascii=False)
        tmp.replace(dst)
    except Exception as exc:
        logger.warning("No se pudo escribir manifest en %s: %s", carpeta_dia, exc)
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass


def _podar(carpeta_sede: Path, retener: int) -> None:
    """Elimina las carpetas diarias más antiguas, conservando solo las últimas `retener`."""
    subdirs = sorted(
        [d for d in carpeta_sede.iterdir() if d.is_dir() and _es_fecha(d.name)],
        key=lambda d: d.name,
    )
    for viejo in subdirs[:-retener]:
        try:
            shutil.rmtree(viejo)
            logger.info("Poda: eliminado respaldo antiguo %s", viejo)
        except Exception as exc:
            logger.warning("No se pudo podar %s: %s", viejo, exc)


# ── Lógica principal de respaldo por sede ────────────────────────────────────

def _respaldar_sede(site: dict, backup_root: Path, hoy: date) -> dict:
    sede = site.get("sede") or site.get("label") or site.get("id", "sin_sede")
    data_dir = Path(site["data_dir"])
    fecha_str = str(hoy)

    resultado: dict = {
        "sede": sede,
        "fecha": fecha_str,
        "archivos_copiados": [],
        "archivos_fallidos": [],
        "valido": False,
        "omitido": False,
        "mensaje": "",
    }

    carpeta_sede = backup_root / sede
    carpeta_dia = carpeta_sede / fecha_str

    # Skip si el respaldo del día ya está completo
    if _backup_dia_completo(carpeta_dia):
        resultado["omitido"] = True
        resultado["valido"] = True
        resultado["mensaje"] = "Respaldo del día ya existe y está completo."
        return resultado

    if not data_dir.exists():
        resultado["mensaje"] = f"Carpeta fuente no existe: {data_dir}"
        return resultado

    archivos = _resolver_archivos_fuente(data_dir)
    if not archivos:
        resultado["mensaje"] = "No se encontraron archivos para respaldar."
        return resultado

    carpeta_dia.mkdir(parents=True, exist_ok=True)

    for src in archivos:
        dst = carpeta_dia / src.name
        # No sobreescribir un archivo bueno existente con una fuente inválida
        if not _validar(src):
            resultado["archivos_fallidos"].append({
                "nombre": src.name,
                "razon": "no pasa validación (archivo corrupto o ilegible)",
            })
            continue
        # Copia atómica: temp → rename
        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                dir=carpeta_dia,
                prefix=src.stem + "_",
                suffix=".tmp",
            ) as tmp_fh:
                tmp_path = Path(tmp_fh.name)
            shutil.copy2(str(src), str(tmp_path))
            tmp_path.replace(dst)
            resultado["archivos_copiados"].append(src.name)
        except Exception as exc:
            try:
                tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
            resultado["archivos_fallidos"].append({"nombre": src.name, "razon": str(exc)})

    resultado["valido"] = (
        bool(resultado["archivos_copiados"]) and not resultado["archivos_fallidos"]
    )
    resultado["mensaje"] = (
        "OK"
        if resultado["valido"]
        else f"{len(resultado['archivos_fallidos'])} archivo(s) no se pudieron respaldar."
    )

    _escribir_manifest(carpeta_dia, {
        "sede": sede,
        "fecha_backup": fecha_str,
        "archivos_copiados": resultado["archivos_copiados"],
        "archivos_fallidos": resultado["archivos_fallidos"],
        "valido": resultado["valido"],
    })

    try:
        _podar(carpeta_sede, RETENTION_DAYS)
    except Exception as exc:
        logger.warning("Error al podar respaldos de %s: %s", sede, exc)

    return resultado


def _escribir_log(backup_root: Path, resultados: list[dict]) -> None:
    log_path = backup_root / "backup_log.jsonl"
    entrada = {
        "timestamp": datetime.now().isoformat(),
        "resultados": resultados,
    }
    try:
        with open(log_path, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entrada, ensure_ascii=False) + "\n")
    except Exception as exc:
        logger.warning("No se pudo escribir en backup_log.jsonl: %s", exc)


# ── Punto de entrada público ──────────────────────────────────────────────────

def ejecutar_backup() -> list[dict]:
    """Recorre todas las sedes remotas y respalda sus archivos críticos.
    Solo opera en super admin con backup_enabled=True y backup_root configurado.
    """
    from app.services.settings_service import get_remote_sites, get_settings, is_super_admin_build

    if not is_super_admin_build():
        return []

    settings = get_settings()
    if not settings.get("backup_enabled"):
        return []

    backup_root_str = str(settings.get("backup_root") or "").strip()
    if not backup_root_str:
        return []

    backup_root = Path(backup_root_str)
    try:
        backup_root.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        logger.error("No se pudo crear la carpeta de respaldo %s: %s", backup_root, exc)
        return []

    hoy = date.today()
    sites = get_remote_sites()
    resultados: list[dict] = []

    for site in sites:
        try:
            res = _respaldar_sede(site, backup_root, hoy)
            resultados.append(res)
            if not res.get("omitido"):
                estado = "OK" if res["valido"] else "FALLO"
                logger.info(
                    "Backup %s [%s]: copiados=%s fallidos=%s",
                    res["sede"], estado,
                    res["archivos_copiados"],
                    [f["nombre"] for f in res["archivos_fallidos"]],
                )
        except Exception as exc:
            logger.error("Error inesperado respaldando sede %s: %s", site.get("sede"), exc)
            resultados.append({
                "sede": site.get("sede") or site.get("id", "?"),
                "valido": False,
                "omitido": False,
                "mensaje": f"Error inesperado: {exc}",
                "archivos_copiados": [],
                "archivos_fallidos": [],
            })

    _escribir_log(backup_root, resultados)
    return resultados


def _run_con_lock() -> None:
    """Wrapper con lock para evitar ejecuciones solapadas."""
    if not _backup_lock.acquire(blocking=False):
        logger.info("Backup ya en ejecución; se omite esta invocación.")
        return
    try:
        ejecutar_backup()
    except Exception as exc:
        logger.error("Error no capturado en backup: %s", exc)
    finally:
        _backup_lock.release()


def _loop_backup(delay_inicial: int) -> None:
    """Espera el delay inicial y luego re-verifica cada REPEAT_SECONDS.
    Cada iteración consulta la configuración en vivo, por lo que activar
    el backup desde Admin en cualquier momento de la sesión surtirá efecto
    en la siguiente pasada del loop sin necesidad de reiniciar.
    """
    import time
    time.sleep(delay_inicial)
    while True:
        _run_con_lock()
        time.sleep(REPEAT_SECONDS)


def programar_backup(delay_segundos: int = DELAY_SECONDS) -> None:
    """Inicia el loop de backup en un hilo daemon. Solo se lanza una vez por proceso."""
    global _backup_scheduled
    with _scheduled_lock:
        if _backup_scheduled:
            return
        _backup_scheduled = True
    t = threading.Thread(target=_loop_backup, args=(delay_segundos,), daemon=True)
    t.start()
    logger.info(
        "Backup automático iniciado (primer intento en %d s, luego cada %d h).",
        delay_segundos, REPEAT_SECONDS // 3600,
    )


def leer_ultimo_log(backup_root_str: str, max_entradas: int = 5) -> list[dict]:
    """Devuelve las últimas entradas del log de respaldos."""
    log_path = Path(backup_root_str) / "backup_log.jsonl"
    if not log_path.exists():
        return []
    lineas: list[str] = []
    try:
        with open(log_path, encoding="utf-8") as fh:
            lineas = fh.readlines()
    except Exception:
        return []
    entradas = []
    for linea in reversed(lineas[-max_entradas * 2:]):
        try:
            entradas.append(json.loads(linea))
            if len(entradas) >= max_entradas:
                break
        except Exception:
            continue
    return entradas


def validar_backup_root(ruta: str) -> dict:
    """Verifica que la carpeta de respaldo exista o se pueda crear y admita escritura."""
    path = Path(str(ruta or "").strip())
    if not ruta or not ruta.strip():
        return {"ok": False, "mensaje": "La ruta está vacía."}
    try:
        path.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        return {"ok": False, "mensaje": f"No se puede crear la carpeta: {exc}"}
    try:
        fd, tmp = tempfile.mkstemp(dir=path, suffix=".tmp")
        import os
        os.close(fd)
        os.unlink(tmp)
    except Exception:
        return {"ok": False, "mensaje": "La carpeta no admite escritura (permisos insuficientes)."}
    return {"ok": True, "mensaje": f"Carpeta válida: {path}"}
